import os
import sys
from openpyxl import load_workbook

infile,sample_name,outdir = sys.argv[1:]

# 打开工作簿
workbook = load_workbook(filename=infile)

# 循环遍历每个工作表
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    print(f"正在处理工作表: {sheet_name}")

    sample_sheet_file = "%s/%s_%s.parsed.xls" % (outdir,sample_name,sheet_name)
    of = open(sample_sheet_file,'w')

    # 遍历当前工作表的行
    for row in sheet.iter_rows(values_only=True):
        print(row)
        row_str = '\t'.join(row)
        of.write(row_str+'\n')